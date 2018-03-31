# -*- coding: utf-8 -*-

"""Adapters module."""

from os.path import exists
from .errors import AdapterError
from openpyxl import load_workbook
import zipfile

import logging
logging.basicConfig(level=logging.INFO)
log = logging.getLogger(__name__)


class ExcelAdapter(object):
    endings = ['xlsx']

    def __init__(self, filepath):
        if not exists(filepath):
            raise AdapterError(f"No such file at {filepath} !!")

        self.filepath = filepath

    def parse(self):
        """Scroll through all sheets in the workbook and return data"""
        data = []
        try:
            wb = load_workbook(filename=self.filepath)
            for sheet in wb.worksheets:
                ws = wb.get_sheet_by_name(sheet.title)
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.value:
                            data.append(str(cell.value))
        except zipfile.BadZipFile as e:
            log.error(e)
        finally:
            return data
